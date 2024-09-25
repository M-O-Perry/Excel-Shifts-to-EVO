import datetime
from PlayActions import send_keys as send

from tkinter import filedialog, simpledialog, messagebox
import os
import openpyxl
from EVOUtil import openTASProgram

def getInputs():
    # get employee ID
    employeeID = simpledialog.askinteger("Employee ID", "Enter the employee ID")
    
    if employeeID is None:
        os._exit(0)
    
    file = filedialog.askopenfilename()
    if file == "":
        os._exit(0)
    
    return employeeID, file

employeeID, file = getInputs()

def getTimes(excelFile):
    wb = openpyxl.load_workbook(excelFile, data_only=True)
    sheet = wb.active

    times = []
    for i in range(3, 20):
        hours = sheet.cell(row = i, column = 24).value
        
        if hours is not None and is_float(hours) and float(hours) > 0:
            op = sheet.cell(row = i, column = 22).value
            sequence = sheet.cell(row = i, column = 23).value
            
            if op is not None and sequence is not None:      
                times.append([op, sequence, hours])
            
    wb.close()
    return times


def is_float(element: any) -> bool:
    #If you expect None to be passed:
    if element is None:
        return False
    try:
        float(element)
        return True
    except ValueError:
        return False
    

def addshift(employeeID, operation, sequence, hours):
    send(["enter", 0.25])
    send(["#"+str(employeeID),  "enter",0.25, "#"+str(operation),  "enter",0.25, "#"+str(sequence), "enter 4", "#"+str(hours), "enter", "alt s", 1])

def addAllShifts(shifts):
    openTASProgram("MDB")
    send(["tab", 1, "right", 0.25, "right", 1, "enter", 4])
    for shift in shifts:
        addshift(employeeID, shift[0], shift[1], shift[2])
    
    send(["alt x", 1])
        
allShifts = getTimes(file)
addAllShifts(allShifts)


messagebox.showinfo("Shifts Added", "Shifts have been added to the system.")
print("Shifts Added", "Shifts have been added to the system.")